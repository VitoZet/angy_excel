import openpyxl
from openpyxl.utils import get_column_letter
from time import time

tic = time()

def ForListDIN_931_933_8_8(x):
    true_name = x.split(sep='М ')
    return true_name[1]


def FullRezba(x):
    full_rezba = x.split(sep='х')
    return full_rezba[0]


def GoodSizeBoltName(bad_size_bolt_name):
    lst_good_name = []
    if '-' in str(bad_size_bolt_name):
        d_x_len = bad_size_bolt_name.split(sep='х')
        len_bolt = d_x_len[1].split(sep='-')
        for i in range(int(len_bolt[0]), int(len_bolt[1]) + 1):
            f = d_x_len[0] + 'х' + str(i)
            lst_good_name.append(f)
    else:
        lst_good_name.append(bad_size_bolt_name)
    return lst_good_name


wb_sale = openpyxl.load_workbook('лю-АПМ БиГ Апрель 80317.xlsx')
ws_sale = wb_sale.get_active_sheet()
########################### БОЛТЫ
wb_bolt = openpyxl.load_workbook('БОЛТ_2017.xlsx')
ws_7798_931_ch = wb_bolt.get_sheet_by_name('7798_931_Ч')
ws_7798_931_zn = wb_bolt.get_sheet_by_name('7798_931_Ц')
ws_7798_8_8 = wb_bolt.get_sheet_by_name('7798_8.8')
ws_DIN__931_933_8_8 = wb_bolt.get_sheet_by_name('DIN 931 933 8.8')
# ws_7798_10_9 = wb_bolt.get_sheet_by_name('7798_10.9')
# ws_7796_8_8 = wb_bolt.get_sheet_by_name('7796_8.8')
# ws_22353_10 = wb_bolt.get_sheet_by_name('22353_10')
# ws_Р52644 = wb_bolt.get_sheet_by_name('Р52644')
# ws_7795 = wb_bolt.get_sheet_by_name('7795')
# ws_7796 = wb_bolt.get_sheet_by_name('7796')
# ws_7801 = wb_bolt.get_sheet_by_name('7801')
# ws_7802 = wb_bolt.get_sheet_by_name('7802')
# ws_7786 = wb_bolt.get_sheet_by_name('7786')

########################## ГАЙКИ
wb_gaika = openpyxl.load_workbook('ГАЙКА_2017.xlsx')
ws_GOST_5915_DIN934_CH = wb_gaika.get_sheet_by_name('ГОСТ_5915_DIN934_Ч')
ws_GOST_5915_DIN934_Zn = wb_gaika.get_sheet_by_name('ГОСТ_5915_DIN934_Ц')
ws_GOST_5915_8 = wb_gaika.get_sheet_by_name('ГОСТ_5915_8')
# ws_22354_110 = wb_gaika.get_sheet_by_name('22354_110')
# ws_Р52645 = wb_gaika.get_sheet_by_name('Р52645')
#######
max_col = ws_sale.max_column
################# Создаем заводы в шапке
ws_sale[get_column_letter(max_col + 1) + '4'] = 'ОСПАЗ (ССМ)'
ws_sale[get_column_letter(max_col + 2) + '4'] = 'ОСПАЗ (ССМ) полная резьба'
ws_sale[get_column_letter(max_col + 3) + '4'] = 'ДМЗ'
ws_sale[get_column_letter(max_col + 4) + '4'] = 'ДМЗ полная резьба'
ws_sale[get_column_letter(max_col + 5) + '4'] = 'ММК'
ws_sale[get_column_letter(max_col + 6) + '4'] = 'ММК полная резьба'
ws_sale[get_column_letter(max_col + 7) + '4'] = 'БелЗан'
ws_sale[get_column_letter(max_col + 8) + '4'] = 'БелЗан (полная резьба)'
ws_sale[get_column_letter(max_col + 9) + '4'] = 'РМЗ'
ws_sale[get_column_letter(max_col + 10) + '4'] = 'РМЗ (полная резьба)'
ws_sale[get_column_letter(max_col + 11) + '4'] = 'ТЕХНОТРОН'
ws_sale[get_column_letter(max_col + 12) + '4'] = 'ТЕХНОТРОН DIN'
ws_sale[get_column_letter(max_col + 13) + '4'] = 'DIN 933'
ws_sale[get_column_letter(max_col + 14) + '4'] = 'КНР'

for nomen_poz in range(5, ws_sale.max_row + 1):
    # name_nomen = ws_sale.cell(row=nomen_poz, column=1).value
    gost = ws_sale.cell(row=nomen_poz, column=15).value
    name_metiz = ws_sale.cell(row=nomen_poz, column=14).value
    coating = ws_sale.cell(row=nomen_poz, column=13).value
    cl_pro4 = ws_sale.cell(row=nomen_poz, column=12).value
    length = ws_sale.cell(row=nomen_poz, column=11).value
    #############'Болт' - 'ГОСТ 7798-70' - 'черный' - 'кл.пр.5.8':
    if name_metiz == 'Болт' and gost == 'ГОСТ 7798-70' and coating == 'черный' and cl_pro4 == 'кл.пр.5.8':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        size = str(diameter) + 'х' + str(length)
        for size_in_marketing in range(7, ws_7798_931_ch.max_row):
            s_i_m = ws_7798_931_ch.cell(row=size_in_marketing, column=2).value
            if 'х' in str(length) and str(diameter) + 'х' + FullRezba(length) in GoodSizeBoltName(s_i_m):
                ws_sale[get_column_letter(max_col + 2) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 11).value
                ws_sale[get_column_letter(max_col + 4) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 9).value
                ws_sale[get_column_letter(max_col + 6) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 7).value
                ws_sale[get_column_letter(max_col + 8) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 5).value
                ws_sale[get_column_letter(max_col + 10) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 3).value
                ws_sale[get_column_letter(max_col + 12) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 1).value
            elif size in GoodSizeBoltName(s_i_m):
                ws_sale[get_column_letter(max_col + 1) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 12).value
                ws_sale[get_column_letter(max_col + 3) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 10).value
                ws_sale[get_column_letter(max_col + 5) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 8).value
                ws_sale[get_column_letter(max_col + 7) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 6).value
                ws_sale[get_column_letter(max_col + 9) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 4).value
                ws_sale[get_column_letter(max_col + 11) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column - 2).value
                ws_sale[get_column_letter(max_col + 13) + str(nomen_poz)] = ws_7798_931_ch.cell(
                    row=size_in_marketing, column=ws_7798_931_ch.max_column).value

                #############'Болт' - 'ГОСТ 7798-70' - 'цинк' - 'кл.пр.5.8':
    elif name_metiz == 'Болт' and gost == 'ГОСТ 7798-70' and coating == 'цинк' and cl_pro4 == 'кл.пр.5.8':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        size = str(diameter) + 'х' + str(length)
        for size_in_marketing in range(7, ws_7798_931_zn.max_row):
            s_i_m = ws_7798_931_ch.cell(row=size_in_marketing, column=2).value
            if 'х' in str(length) and str(diameter) + 'х' + FullRezba(length) in GoodSizeBoltName(s_i_m):
                ws_sale[get_column_letter(max_col + 2) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 11).value
                ws_sale[get_column_letter(max_col + 4) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 9).value
                ws_sale[get_column_letter(max_col + 6) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 7).value
                ws_sale[get_column_letter(max_col + 8) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 5).value
                ws_sale[get_column_letter(max_col + 10) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                                column=ws_7798_931_zn.max_column - 3).value
                ws_sale[get_column_letter(max_col + 12) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                                column=ws_7798_931_zn.max_column - 1).value
            elif size in GoodSizeBoltName(s_i_m):
                ws_sale[get_column_letter(max_col + 1) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 12).value
                ws_sale[get_column_letter(max_col + 3) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 10).value
                ws_sale[get_column_letter(max_col + 5) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 8).value
                ws_sale[get_column_letter(max_col + 7) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 6).value
                ws_sale[get_column_letter(max_col + 9) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                               column=ws_7798_931_zn.max_column - 4).value
                ws_sale[get_column_letter(max_col + 11) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                                column=ws_7798_931_zn.max_column - 2).value
                ws_sale[get_column_letter(max_col + 13) + str(nomen_poz)] = ws_7798_931_zn.cell(row=size_in_marketing,
                                                                                                column=ws_7798_931_zn.max_column).value
                ############'Болт'-'ГОСТ 7798-70'-'черный'-'кл.пр.8.8':
    elif name_metiz == 'Болт' and gost == 'ГОСТ 7798-70' and coating == 'черный' and cl_pro4 == 'кл.пр.8.8':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        size = str(diameter) + 'х' + str(length)
        for size_in_marketing in range(5, ws_7798_8_8.max_row):
            s_i_m = ws_7798_8_8.cell(row=size_in_marketing, column=2).value
            if 'х' in str(length) and str(diameter) + 'х' + FullRezba(length) in GoodSizeBoltName(
                    s_i_m):  # полная резьба
                ws_sale[get_column_letter(max_col + 4) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 10).value  # ДМЗ ПОЛНАЯ
                ws_sale[get_column_letter(max_col + 10) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                             column=ws_7798_8_8.max_column - 8).value  # РМЗ ПОЛНАЯ
            elif size in GoodSizeBoltName(s_i_m):
                ws_sale[get_column_letter(max_col + 1) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 13).value
                ws_sale[get_column_letter(max_col + 5) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 12).value
                ws_sale[get_column_letter(max_col + 3) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 11).value
                ws_sale[get_column_letter(max_col + 9) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 9).value
                ws_sale[get_column_letter(max_col + 11) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                             column=ws_7798_8_8.max_column - 7).value
        for size_in_marketing in range(4, ws_DIN__931_933_8_8.max_row):
            s_i_m = ws_DIN__931_933_8_8.cell(row=size_in_marketing, column=2).value
            if 'х' in str(length) and str(diameter) + 'х' + FullRezba(length) in ForListDIN_931_933_8_8(s_i_m):  # полная резьба
                ws_sale[get_column_letter(max_col + 8) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(row=size_in_marketing,column=ws_DIN__931_933_8_8.max_column - 5).value #БЕЛЗАН (DIN 933)
                ws_sale[get_column_letter(max_col + 13) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(row=size_in_marketing,column=ws_DIN__931_933_8_8.max_column - 4).value #ТЕХНОТРОН DIN 931/933
            elif size in ForListDIN_931_933_8_8(s_i_m): #обычная резьба
                ws_sale[get_column_letter(max_col + 7) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(row=size_in_marketing,column=ws_DIN__931_933_8_8.max_column - 6).value #БЕЛЗАН (DIN 931)
                ws_sale[get_column_letter(max_col + 13) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(row=size_in_marketing,column=ws_DIN__931_933_8_8.max_column - 4).value #ТЕХНОТРОН DIN 931/933
##############'Болт'--'ГОСТ 7798-70'--'цинк'--'кл.пр.8.8':
    elif name_metiz == 'Болт' and gost == 'ГОСТ 7798-70' and coating == 'цинк' and cl_pro4 == 'кл.пр.8.8':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        size = str(diameter) + 'х' + str(length)
        for size_in_marketing in range(5, ws_7798_8_8.max_row):
            s_i_m = ws_7798_8_8.cell(row=size_in_marketing, column=2).value
            if 'х' in str(length) and str(diameter) + 'х' + FullRezba(length) in GoodSizeBoltName(
                    s_i_m):  # полная резьба
                ws_sale[get_column_letter(max_col + 4) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 3).value  # ДМЗ ПОЛНАЯ
                ws_sale[get_column_letter(max_col + 10) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                             column=ws_7798_8_8.max_column - 1).value  # РМЗ ПОЛНАЯ
            elif size in GoodSizeBoltName(s_i_m):
                ws_sale[get_column_letter(max_col + 1) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 6).value
                ws_sale[get_column_letter(max_col + 5) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 5).value
                ws_sale[get_column_letter(max_col + 3) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 4).value
                ws_sale[get_column_letter(max_col + 9) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                            column=ws_7798_8_8.max_column - 2).value
                ws_sale[get_column_letter(max_col + 11) + str(nomen_poz)] = ws_7798_8_8.cell(row=size_in_marketing,
                                                                                             column=ws_7798_8_8.max_column).value
        for size_in_marketing in range(4, ws_DIN__931_933_8_8.max_row):
            s_i_m = ws_DIN__931_933_8_8.cell(row=size_in_marketing, column=2).value
            if 'х' in str(length) and str(diameter) + 'х' + FullRezba(length) in ForListDIN_931_933_8_8(
                    s_i_m):  # полная резьба
                ws_sale[get_column_letter(max_col + 8) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(
                    row=size_in_marketing, column=ws_DIN__931_933_8_8.max_column - 1).value  # БЕЛЗАН (DIN 933)
                ws_sale[get_column_letter(max_col + 13) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(
                    row=size_in_marketing, column=ws_DIN__931_933_8_8.max_column).value  # ТЕХНОТРОН DIN 931/933
            elif size in ForListDIN_931_933_8_8(s_i_m):  # обычная резьба
                ws_sale[get_column_letter(max_col + 7) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(
                    row=size_in_marketing, column=ws_DIN__931_933_8_8.max_column - 2).value  # БЕЛЗАН (DIN 931)
                ws_sale[get_column_letter(max_col + 13) + str(nomen_poz)] = ws_DIN__931_933_8_8.cell(
                    row=size_in_marketing, column=ws_DIN__931_933_8_8.max_column).value  # ТЕХНОТРОН DIN 931/933
    # ГАЙКА -- ГОСТ 5915-70 -- черный -- кл.пр.6
    elif name_metiz == 'Гайка' and gost == 'ГОСТ 5915-70' and coating == 'черный' and cl_pro4 == 'кл.пр.6':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        for size_in_marketing in range(6, ws_GOST_5915_DIN934_CH.max_row):
            d_in_mrktng = ws_GOST_5915_DIN934_CH.cell(row=size_in_marketing,column=2).value
            if str(diameter) == str(d_in_mrktng):
                ws_sale[get_column_letter(max_col + 1)+ str(nomen_poz)]=ws_GOST_5915_DIN934_CH.cell(row=size_in_marketing, column=ws_GOST_5915_DIN934_CH.max_column -5).value #ОСПАЗ (ССМ)
                ws_sale[get_column_letter(max_col + 3)+ str(nomen_poz)]=ws_GOST_5915_DIN934_CH.cell(row=size_in_marketing, column=ws_GOST_5915_DIN934_CH.max_column -4).value #ДМЗ
                ws_sale[get_column_letter(max_col + 5)+ str(nomen_poz)]=ws_GOST_5915_DIN934_CH.cell(row=size_in_marketing, column=ws_GOST_5915_DIN934_CH.max_column -3).value #ММК
                ws_sale[get_column_letter(max_col + 9)+ str(nomen_poz)]=ws_GOST_5915_DIN934_CH.cell(row=size_in_marketing, column=ws_GOST_5915_DIN934_CH.max_column -2).value #РМЗ
                ws_sale[get_column_letter(max_col + 11)+ str(nomen_poz)]=ws_GOST_5915_DIN934_CH.cell(row=size_in_marketing, column=ws_GOST_5915_DIN934_CH.max_column -1).value #ТЕХНОТРОН
                ws_sale[get_column_letter(max_col + 14)+ str(nomen_poz)]=ws_GOST_5915_DIN934_CH.cell(row=size_in_marketing, column=ws_GOST_5915_DIN934_CH.max_column).value #КНР DIN 934
    # ГАЙКА -- ГОСТ 5915-70 -- цинк -- кл.пр.6
    elif name_metiz == 'Гайка' and gost == 'ГОСТ 5915-70' and coating == 'цинк' and cl_pro4 == 'кл.пр.6':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        for size_in_marketing in range(5, ws_GOST_5915_DIN934_Zn.max_row):
            d_in_mrktng = ws_GOST_5915_DIN934_Zn.cell(row=size_in_marketing, column=1).value
            if str(diameter) == str(d_in_mrktng):
                            ws_sale[get_column_letter(max_col + 1) + str(nomen_poz)] = ws_GOST_5915_DIN934_Zn.cell(
                                row=size_in_marketing, column=ws_GOST_5915_DIN934_Zn.max_column - 5).value  # ОСПАЗ (ССМ)
                            ws_sale[get_column_letter(max_col + 3) + str(nomen_poz)] = ws_GOST_5915_DIN934_Zn.cell(
                                row=size_in_marketing, column=ws_GOST_5915_DIN934_Zn.max_column - 4).value  # ДМЗ
                            ws_sale[get_column_letter(max_col + 5) + str(nomen_poz)] = ws_GOST_5915_DIN934_Zn.cell(
                                row=size_in_marketing, column=ws_GOST_5915_DIN934_Zn.max_column - 3).value  # ММК
                            ws_sale[get_column_letter(max_col + 9) + str(nomen_poz)] = ws_GOST_5915_DIN934_Zn.cell(
                                row=size_in_marketing, column=ws_GOST_5915_DIN934_Zn.max_column - 2).value  # РМЗ
                            ws_sale[get_column_letter(max_col + 11) + str(nomen_poz)] = ws_GOST_5915_DIN934_Zn.cell(
                                row=size_in_marketing, column=ws_GOST_5915_DIN934_Zn.max_column - 1).value  # ТЕХНОТРОН
                            ws_sale[get_column_letter(max_col + 14) + str(nomen_poz)] = ws_GOST_5915_DIN934_Zn.cell(
                                row=size_in_marketing, column=ws_GOST_5915_DIN934_Zn.max_column).value  # КНР DIN 934
    # ГАЙКА -- ГОСТ 5915-70 -- черный -- кл.пр.8
    elif name_metiz == 'Гайка' and gost == 'ГОСТ 5915-70' and coating == 'черный' and cl_pro4 == 'кл.пр.8':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        for size_in_marketing in range(7, ws_GOST_5915_8.max_row):
            d_in_mrktng = ws_GOST_5915_8.cell(row=size_in_marketing,column=2).value
            if str(diameter) == str(d_in_mrktng).replace('М ' or 'М', ''):
                ws_sale[get_column_letter(max_col + 1)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -9).value #ОСПАЗ (ССМ)
                ws_sale[get_column_letter(max_col + 3)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -8).value #ДМЗ
                ws_sale[get_column_letter(max_col + 5)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -7).value #ММК
                ws_sale[get_column_letter(max_col + 9)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -6).value #РМЗ
                ws_sale[get_column_letter(max_col + 11)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -5).value #ТЕХНОТРОН
    # ГАЙКА -- ГОСТ 5915-70 -- цинк -- кл.пр.8
    elif name_metiz == 'Гайка' and gost == 'ГОСТ 5915-70' and coating == 'цинк' and cl_pro4 == 'кл.пр.8':
        diameter = ws_sale.cell(row=nomen_poz, column=10).value.replace('М' or '2M' or '3M', '')
        for size_in_marketing in range(7, ws_GOST_5915_8.max_row):
            d_in_mrktng = ws_GOST_5915_8.cell(row=size_in_marketing,column=2).value
            if str(diameter) == str(d_in_mrktng).replace('М ' or 'М', ''):
                ws_sale[get_column_letter(max_col + 1)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -4).value #ОСПАЗ (ССМ)
                ws_sale[get_column_letter(max_col + 3)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -3).value #ДМЗ
                ws_sale[get_column_letter(max_col + 5)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -2).value #ММК
                ws_sale[get_column_letter(max_col + 9)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column -1).value #РМЗ
                ws_sale[get_column_letter(max_col + 11)+ str(nomen_poz)]=ws_GOST_5915_8.cell(row=size_in_marketing, column=ws_GOST_5915_8.max_column).value #ТЕХНОТРОН

    print(nomen_poz)
wb_sale.save('PRICE_Angy.xlsx')
print('сохраняю...')
toc = time()
print(toc - tic)
print('Готово, проверяй.')