import openpyxl.cell
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

wb = openpyxl.load_workbook('АПМ БИ Г - 13 мес.xlsx')
ws = wb.get_active_sheet()
head = ws['4']
# fill_red = PatternFill(start_color='F44336', fill_type='solid')
fill_red2 = PatternFill(start_color='E57373', fill_type='solid')
fill_min = PatternFill(start_color='3D5AFE', fill_type='solid')
fill_min2 = PatternFill(start_color='8C9EFF', fill_type='solid')
fill_except = PatternFill(start_color='9C27B0', fill_type='solid')
fill_more5 = PatternFill(start_color='FFC400', fill_type='solid')  # от 5.5
fill_3btw5 = PatternFill(start_color='E040FB', fill_type='solid')  #

ws['B1'] = 'Средн<3'
ws['B1'].fill = fill_min
ws['C1'] = 'Средн<2'
ws['C1'].fill = fill_min2
ws['D1'] = '2<Ср<3'
ws['D1'].fill = fill_red2
ws['E1'] = '3<Ср<5.5'
ws['E1'].fill = fill_3btw5
ws['F1'] = 'Ср>5.5'
ws['F1'].fill = fill_more5


def SearchMedium():
    for med in head:
        if med.value == 'Среднее':
            return column_index_from_string(med.column)


def SearchLastDate():
    for ld in head:
        if ld.value == 'Итого':
            return column_index_from_string(ld.column)  # Значение столбцов привел к ИНТ и минусанул 1


er = []
for rows in range(5, ws.max_row):
    for col in range(18, SearchLastDate()):
        x = ws.cell(row=rows, column=col)
        y = ws.cell(row=rows, column=SearchMedium())
        try:
            if x.value / y.value >= 5.5:
                ja4 = ws[x.coordinate]
                ja4.fill = fill_more5
            elif 3 <= x.value / y.value < 5.5:
                ja4 = ws[x.coordinate]
                ja4.fill = fill_4btw5
            # elif 2 <= x.value / y.value < 3:
            #     ja4 = ws[x.coordinate]
            #     ja4.fill = fill_red
            elif x.value / y.value > 2:
                ja4 = ws[x.coordinate]
                ja4.fill = fill_red2
            elif x.value / y.value < 0.33:
                ja4 = ws[x.coordinate]
                ja4.fill = fill_min
            elif x.value / y.value < 0.5:
                ja4 = ws[x.coordinate]
                ja4.fill = fill_min2

        except:
            if x == None:
                ja4 = [x.coordinate]
                ja4.fill = fill_except

wb.save('copy_angy_test.xlsx')
