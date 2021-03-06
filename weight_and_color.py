import openpyxl
import openpyxl.cell
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles.fonts import Font
from openpyxl.styles import PatternFill

wb_sale = openpyxl.load_workbook('АПМ-o1.xlsx')
ws_sale = wb_sale.get_active_sheet()
wb_sale.create_sheet(title='Диам в тонн')
ws_weight = wb_sale.get_sheet_by_name('Диам в тонн')
wb_sale.create_sheet(title='ТАБЛ 1')
ws_tabl1 = wb_sale.get_sheet_by_name('ТАБЛ 1')
wb_sale.create_sheet(title='ТАБЛ 2')
ws_tabl2 = wb_sale.get_sheet_by_name('ТАБЛ 2')

fill_ALL = PatternFill(start_color='DDDDDD', fill_type='solid')
fill_ALL_Bolt = PatternFill(start_color='8FBC8F', fill_type='solid')
fill_ALL_Gaika = PatternFill(start_color='66CDAA', fill_type='solid')
fill_ITOGO = PatternFill(start_color='DDA0DD', fill_type='solid')
fill_MOST = PatternFill(start_color='FFA500', fill_type='solid')
head_style = Font(b=True)


def HeadWrite(lst_excel):
    lst_excel['A1'] = ws_sale['I4'].value
    lst_excel['A1'].font = head_style
    lst_excel['B1'] = ws_sale['N4'].value
    lst_excel['B1'].font = head_style
    lst_excel['C1'] = ws_sale['M4'].value
    lst_excel['C1'].font = head_style
    lst_excel['D1'] = ws_sale['L4'].value
    lst_excel['D1'].font = head_style
    lst_excel['E1'] = ws_sale['O4'].value
    lst_excel['E1'].font = head_style
    lst_excel['F1'] = ws_sale['J4'].value
    lst_excel['F1'].font = head_style
    for e, mm in enumerate(lst_sale):
        lst_excel[get_column_letter(e + 7) + '1'] = mm
        lst_excel[get_column_letter(e + 7) + '1'].font = head_style


def SearchLastDate():
    for ld in ws_sale['4']:
        if ld.value == 'Склад':
            return column_index_from_string(ld.column)


lst_sale = []

tonnageData = {}
# Сохраняем таблицу в словарь
for m_s in range(18, SearchLastDate()):
    lst_sale.append(ws_sale.cell(row=4, column=m_s).value)
    for nomen_poz in range(5, ws_sale.max_row + 1):
        ### name_nomen = ws_sale.cell(row=nomen_poz, column=1).value
        gost = ws_sale.cell(row=nomen_poz, column=15).value
        name_metiz = ws_sale.cell(row=nomen_poz, column=14).value
        coating = ws_sale.cell(row=nomen_poz, column=13).value
        cl_pro4 = ws_sale.cell(row=nomen_poz, column=12).value
        ###    length = ws_sale.cell(row=nomen_poz, column=11).value
        diameter = ws_sale.cell(row=nomen_poz, column=10).value  # .replace('2M' or '3M', 'М')
        measure_unit = ws_sale.cell(row=nomen_poz, column=6).value
        sklad = ws_sale.cell(row=nomen_poz, column=9).value
        month_sale = ws_sale.cell(row=4, column=m_s).value
        weight = ws_sale.cell(row=nomen_poz, column=m_s).value
        if measure_unit and name_metiz and coating and cl_pro4 and diameter and weight:
            tonnageData.setdefault(sklad, {})
            tonnageData[sklad].setdefault(measure_unit, {})
            tonnageData[sklad][measure_unit].setdefault(name_metiz, {})
            tonnageData[sklad][measure_unit][name_metiz].setdefault(coating, {})
            tonnageData[sklad][measure_unit][name_metiz][coating].setdefault(cl_pro4, {})
            tonnageData[sklad][measure_unit][name_metiz][coating][cl_pro4].setdefault(gost, {})
            tonnageData[sklad][measure_unit][name_metiz][coating][cl_pro4][gost].setdefault(diameter, {})
            tonnageData[sklad][measure_unit][name_metiz][coating][cl_pro4][gost][diameter].setdefault(month_sale,
                                                                                                      {'weight': 0})
            tonnageData[sklad][measure_unit][name_metiz][coating][cl_pro4][gost][diameter][month_sale][
                'weight'] += float(weight)

# Словарь с списками необходимых группировок диаметров
DictGroupDiam = {'М16-М30': ('М16', 'М18', 'М20', 'М22', 'М24', 'М27', 'М30'),
                 'М6-М16': ('М6', 'М8', 'М10', 'М12', 'М14', 'М16'),
                 'М18-М36': ('М18', 'М20', 'М22', 'М24', 'М27', 'М30', 'М36', 'М33'),
                 'М4-М16': ('М4', 'М5', 'М6', 'М8', 'М10', 'М12', 'М14', 'М16'),
                 'М3, М42-М72': ('М3', 'М42', 'М45', 'М48', 'М52', 'М56', 'М64', 'М72')}


def SumGroupDiamMonth(groupDiam, sklad, name_metiz, coating, cl_pro4, gost, month):
    '''Сумма веса в указанном диапозоне диаметров по указанному месяцу(месяц дает другая функция)'''
    summa = 0
    for d in groupDiam:
        try:
            summa += tonnageData[sklad]['кг'][name_metiz][coating][cl_pro4][gost][d][month]['weight'] / 1000
        except:
            summa += 0
    return summa


def TAB1SumGroupDiamMonthForCol(groupDiam, sklad, name_metiz, coating, cl_pro4, gost):
    '''Сохранение в Excel по коллонкам суммы весов по месяцам'''
    for e, month_col in enumerate(lst_sale):
        if coating == 'ч + ц':
            ws_tabl1[get_column_letter(e + 7) + str(ws_tabl1.max_row)] = round(
                SumGroupDiamMonth(groupDiam, sklad, name_metiz, 'черный', cl_pro4, gost, month_col), 5) + round(
                SumGroupDiamMonth(groupDiam, sklad, name_metiz, 'цинк', cl_pro4, gost, month_col), 5)
        else:
            ws_tabl1[get_column_letter(e + 7) + str(ws_tabl1.max_row)] = round(
                SumGroupDiamMonth(groupDiam, sklad, name_metiz, coating, cl_pro4, gost, month_col), 5)


def TAB2SumGroupDiamMonthForCol(groupDiam, name_metiz, cl_pro4, gost):
    '''Сохранение в Excel по коллонкам в ТАБ2'''
    coating = ('черный', 'цинк', 'ч+ц')
    for co in coating:
        if co != 'ч+ц':
            ws_tabl2['A' + str(ws_tabl2.max_row + 1)] = 'S+SZ+Z'
            ws_tabl2['B' + str(ws_tabl2.max_row)] = name_metiz
            ws_tabl2['C' + str(ws_tabl2.max_row)] = co
            ws_tabl2['D' + str(ws_tabl2.max_row)] = cl_pro4
            ws_tabl2['E' + str(ws_tabl2.max_row)] = gost
            ws_tabl2['F' + str(ws_tabl2.max_row)] = groupDiam
            for e, month_col in enumerate(lst_sale):
                ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)] = round(
                    SumGroupDiamMonth(DictGroupDiam[groupDiam], 'S', name_metiz, co, cl_pro4, gost, month_col),
                    5) + round(
                    SumGroupDiamMonth(DictGroupDiam[groupDiam], 'SZ', name_metiz, co, cl_pro4, gost, month_col),
                    5) + round(
                    SumGroupDiamMonth(DictGroupDiam[groupDiam], 'Z', name_metiz, co, cl_pro4, gost, month_col), 5)
        else:
            ws_tabl2['A' + str(ws_tabl2.max_row + 1)] = 'ALL'
            ws_tabl2['A' + str(ws_tabl2.max_row)].fill = fill_ALL
            ws_tabl2['B' + str(ws_tabl2.max_row)] = name_metiz
            ws_tabl2['B' + str(ws_tabl2.max_row)].fill = fill_ALL
            ws_tabl2['C' + str(ws_tabl2.max_row)] = co
            ws_tabl2['C' + str(ws_tabl2.max_row)].fill = fill_ALL
            ws_tabl2['C' + str(ws_tabl2.max_row)].font = head_style
            ws_tabl2['D' + str(ws_tabl2.max_row)] = cl_pro4
            ws_tabl2['D' + str(ws_tabl2.max_row)].fill = fill_ALL
            ws_tabl2['E' + str(ws_tabl2.max_row)] = gost
            ws_tabl2['E' + str(ws_tabl2.max_row)].fill = fill_ALL
            ws_tabl2['F' + str(ws_tabl2.max_row)] = groupDiam
            ws_tabl2['F' + str(ws_tabl2.max_row)].fill = fill_ALL
            for e, month_col in enumerate(lst_sale):
                ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)] = ws_tabl2[get_column_letter(e + 7) + str(
                    ws_tabl2.max_row - 1)].value + ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row - 2)].value
                ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)].fill = fill_ALL


def TAB2SumBolT(cl_pro4, gost):
    '''Сохранение в Excel cумму общих болтов по коллонкам в ТАБ2'''
    coating = ('черный', 'цинк', 'ч+ц')
    for co in coating:
        ws_tabl2['A' + str(ws_tabl2.max_row + 1)] = 'All Diam'
        ws_tabl2['A' + str(ws_tabl2.max_row)].fill = fill_ALL_Bolt
        ws_tabl2['B' + str(ws_tabl2.max_row)] = 'Болт'
        ws_tabl2['B' + str(ws_tabl2.max_row)].fill = fill_ALL_Bolt
        ws_tabl2['C' + str(ws_tabl2.max_row)] = co
        ws_tabl2['C' + str(ws_tabl2.max_row)].fill = fill_ALL_Bolt
        ws_tabl2['D' + str(ws_tabl2.max_row)] = cl_pro4
        ws_tabl2['D' + str(ws_tabl2.max_row)].fill = fill_ALL_Bolt
        ws_tabl2['E' + str(ws_tabl2.max_row)] = gost
        ws_tabl2['E' + str(ws_tabl2.max_row)].fill = fill_ALL_Bolt
        ws_tabl2['F' + str(ws_tabl2.max_row)] = 'М6-М36'
        ws_tabl2['F' + str(ws_tabl2.max_row)].fill = fill_ALL_Bolt
        ws_tabl2['F' + str(ws_tabl2.max_row)].font = head_style
        for e, month_col in enumerate(lst_sale):
            ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)] = ws_tabl2[get_column_letter(e + 7) + str(
                ws_tabl2.max_row - 3)].value + ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row - 6)].value
            ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)].fill = fill_ALL_Bolt
        if co == 'ч+ц':
            ws_tabl2['C' + str(ws_tabl2.max_row)].font = head_style


def TAB2SumGaika(cl_pro4, gost):
    '''Сохранение в Excel cумму общих гаек по коллонкам в ТАБ2'''
    coating = ('черный', 'цинк', 'ч+ц')
    for co in coating:
        ws_tabl2['A' + str(ws_tabl2.max_row + 1)] = 'All Diam'
        ws_tabl2['A' + str(ws_tabl2.max_row)].fill = fill_ALL_Gaika
        ws_tabl2['B' + str(ws_tabl2.max_row)] = 'Гайка'
        ws_tabl2['B' + str(ws_tabl2.max_row)].fill = fill_ALL_Gaika
        ws_tabl2['C' + str(ws_tabl2.max_row)] = co
        ws_tabl2['C' + str(ws_tabl2.max_row)].fill = fill_ALL_Gaika
        ws_tabl2['D' + str(ws_tabl2.max_row)] = cl_pro4
        ws_tabl2['D' + str(ws_tabl2.max_row)].fill = fill_ALL_Gaika
        ws_tabl2['E' + str(ws_tabl2.max_row)] = gost
        ws_tabl2['E' + str(ws_tabl2.max_row)].fill = fill_ALL_Gaika
        ws_tabl2['F' + str(ws_tabl2.max_row)] = 'М3-М72'
        ws_tabl2['F' + str(ws_tabl2.max_row)].fill = fill_ALL_Gaika
        ws_tabl2['F' + str(ws_tabl2.max_row)].font = head_style
        for e, month_col in enumerate(lst_sale):
            ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)] = ws_tabl2[get_column_letter(e + 7) + str(
                ws_tabl2.max_row - 3)].value + ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row - 6)].value + \
                                                                         ws_tabl2[get_column_letter(e + 7) + str(
                                                                             ws_tabl2.max_row - 9)].value
            ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)].fill = fill_ALL_Gaika
        if co == 'ч+ц':
            ws_tabl2['C' + str(ws_tabl2.max_row)].font = head_style


def TAB1AllSumGroupDiamMonthForCol():
    '''Считаем сумму по всем скалдам и записываем по всем столбцам'''
    for e, month_col in enumerate(lst_sale):
        ws_tabl1[get_column_letter(e + 7) + str(ws_tabl1.max_row)] = ws_tabl1[get_column_letter(e + 7) + str(
            ws_tabl1.max_row - 1)].value + ws_tabl1[get_column_letter(e + 7) + str(ws_tabl1.max_row - 2)].value + \
                                                                     ws_tabl1[get_column_letter(e + 7) + str(
                                                                         ws_tabl1.max_row - 3)].value
        ws_tabl1[get_column_letter(e + 7) + str(ws_tabl1.max_row)].fill = fill_ALL


def TAB1SaveInExcel(name_metiz, coating, cl_pro4, gost, groupDiam):
    lst_sklad = ['S', 'Z', 'SZ', 'ALL']
    for sk in lst_sklad:
        ws_tabl1['A' + str(ws_tabl1.max_row + 1)] = sk
        ws_tabl1['B' + str(ws_tabl1.max_row)] = name_metiz
        ws_tabl1['C' + str(ws_tabl1.max_row)] = coating
        ws_tabl1['D' + str(ws_tabl1.max_row)] = cl_pro4
        ws_tabl1['E' + str(ws_tabl1.max_row)] = gost
        ws_tabl1['F' + str(ws_tabl1.max_row)] = groupDiam
        if not sk == 'ALL':
            TAB1SumGroupDiamMonthForCol(DictGroupDiam[groupDiam], sk, name_metiz, coating, cl_pro4, gost)
        else:
            TAB1AllSumGroupDiamMonthForCol()
            ws_tabl1['A' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['B' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['C' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['D' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['E' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['F' + str(ws_tabl1.max_row)].fill = fill_ALL


def TAB1SaveInExcelSummCoating(name_metiz, coating, cl_pro4, gost, groupDiam):
    lst_sklad = ['S', 'Z', 'SZ', 'ALL']
    for sk in lst_sklad:
        ws_tabl1['A' + str(ws_tabl1.max_row + 1)] = sk
        ws_tabl1['B' + str(ws_tabl1.max_row)] = name_metiz
        ws_tabl1['C' + str(ws_tabl1.max_row)] = 'ч + ц'
        ws_tabl1['D' + str(ws_tabl1.max_row)] = cl_pro4
        ws_tabl1['E' + str(ws_tabl1.max_row)] = gost
        ws_tabl1['F' + str(ws_tabl1.max_row)] = groupDiam
        if not sk == 'ALL':
            TAB1SumGroupDiamMonthForCol(DictGroupDiam[groupDiam], sk, name_metiz, coating, cl_pro4, gost)
        else:
            TAB1AllSumGroupDiamMonthForCol()
            ws_tabl1['A' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['B' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['C' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['C' + str(ws_tabl1.max_row)].font = head_style
            ws_tabl1['D' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['E' + str(ws_tabl1.max_row)].fill = fill_ALL
            ws_tabl1['F' + str(ws_tabl1.max_row)].fill = fill_ALL


def ItogoBolt():
    ws_tabl2['A' + str(ws_tabl2.max_row + 1)] = 'ИТОГО'
    ws_tabl2['A' + str(ws_tabl2.max_row)].font = head_style
    ws_tabl2['B' + str(ws_tabl2.max_row)] = 'БОЛТЫ'
    ws_tabl2['B' + str(ws_tabl2.max_row)].font = head_style
    ws_tabl2['E' + str(ws_tabl2.max_row)] = 'ГОСТ 7798-70'
    ws_tabl2['E' + str(ws_tabl2.max_row)].font = head_style
    for e, month_col in enumerate(lst_sale):
        ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)] = ws_tabl2[get_column_letter(e + 7) + str(
            ws_tabl2.max_row - 1)].value + ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row - 10)].value
    for i in range(1, ws_tabl2.max_column + 1):
        ws_tabl2[get_column_letter(i) + str(ws_tabl2.max_row)].fill = fill_ITOGO


def ItogoGaika():
    ws_tabl2['A' + str(ws_tabl2.max_row + 1)] = 'ИТОГО'
    ws_tabl2['A' + str(ws_tabl2.max_row)].font = head_style
    ws_tabl2['B' + str(ws_tabl2.max_row)] = 'ГАЙКИ'
    ws_tabl2['B' + str(ws_tabl2.max_row)].font = head_style
    ws_tabl2['E' + str(ws_tabl2.max_row)] = 'ГОСТ 5915-70'
    ws_tabl2['E' + str(ws_tabl2.max_row)].font = head_style
    for e, month_col in enumerate(lst_sale):
        ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)] = ws_tabl2[get_column_letter(e + 7) + str(
            ws_tabl2.max_row - 1)].value + ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row - 13)].value
    for i in range(1, ws_tabl2.max_column + 1):
        ws_tabl2[get_column_letter(i) + str(ws_tabl2.max_row)].fill = fill_ITOGO


def TAB2MostKrep(name_metiz, cl_pro4, gost, groupDiam):
    ws_tabl2['A' + str(ws_tabl2.max_row + 1)] = 'S+SZ+Z'
    ws_tabl2['B' + str(ws_tabl2.max_row)] = name_metiz
    ws_tabl2['C' + str(ws_tabl2.max_row)] = 'ч + ц'
    ws_tabl2['D' + str(ws_tabl2.max_row)] = cl_pro4
    ws_tabl2['E' + str(ws_tabl2.max_row)] = gost
    ws_tabl2['F' + str(ws_tabl2.max_row)] = groupDiam
    sklad = ('S', 'SZ', 'Z')
    coating = ('черный', 'цинк')
    for e, month_col in enumerate(lst_sale):
        for sk in sklad:
            for co in coating:
                summ = 0
                summ += SumGroupDiamMonth(DictGroupDiam[groupDiam], sk, name_metiz, co, cl_pro4, gost, month_col)
                ws_tabl2[get_column_letter(e + 7) + str(ws_tabl2.max_row)] = summ
    for i in range(1, ws_tabl2.max_column + 1):
        ws_tabl2[get_column_letter(i) + str(ws_tabl2.max_row)].fill = fill_MOST


# '''Пишем Болты в ТАБ1'''
TAB1SaveInExcelSummCoating('Болт', 'ч + ц', 'кл.пр.10.9', 'ГОСТ Р 52644-2006', 'М16-М30')
TAB1SaveInExcel('Болт', 'черный', 'кл.пр.5.8', 'ГОСТ 7798-70', 'М6-М16')
TAB1SaveInExcel('Болт', 'черный', 'кл.пр.8.8', 'ГОСТ 7798-70', 'М6-М16')
TAB1SaveInExcel('Болт', 'цинк', 'кл.пр.5.8', 'ГОСТ 7798-70', 'М6-М16')
TAB1SaveInExcel('Болт', 'цинк', 'кл.пр.8.8', 'ГОСТ 7798-70', 'М6-М16')
TAB1SaveInExcel('Болт', 'черный', 'кл.пр.5.8', 'ГОСТ 7798-70', 'М18-М36')
TAB1SaveInExcel('Болт', 'черный', 'кл.пр.8.8', 'ГОСТ 7798-70', 'М18-М36')
TAB1SaveInExcel('Болт', 'цинк', 'кл.пр.5.8', 'ГОСТ 7798-70', 'М18-М36')
TAB1SaveInExcel('Болт', 'цинк', 'кл.пр.8.8', 'ГОСТ 7798-70', 'М18-М36')
# Пишем Гайки в ТАБ1
TAB1SaveInExcelSummCoating('Гайка', 'ч + ц', 'кл.пр.10', 'ГОСТ Р 52645-2006', 'М16-М30')
TAB1SaveInExcel('Гайка', 'черный', 'кл.пр.6', 'ГОСТ 5915-70', 'М4-М16')
TAB1SaveInExcel('Гайка', 'черный', 'кл.пр.6', 'ГОСТ 5915-70', 'М18-М36')
TAB1SaveInExcel('Гайка', 'черный', 'кл.пр.6', 'ГОСТ 5915-70', 'М3, М42-М72')

TAB1SaveInExcel('Гайка', 'черный', 'кл.пр.8', 'ГОСТ 5915-70', 'М4-М16')
TAB1SaveInExcel('Гайка', 'черный', 'кл.пр.8', 'ГОСТ 5915-70', 'М18-М36')
TAB1SaveInExcel('Гайка', 'черный', 'кл.пр.8', 'ГОСТ 5915-70', 'М3, М42-М72')

TAB1SaveInExcel('Гайка', 'цинк', 'кл.пр.6', 'ГОСТ 5915-70', 'М4-М16')
TAB1SaveInExcel('Гайка', 'цинк', 'кл.пр.6', 'ГОСТ 5915-70', 'М18-М36')
TAB1SaveInExcel('Гайка', 'цинк', 'кл.пр.6', 'ГОСТ 5915-70', 'М3, М42-М72')

TAB1SaveInExcel('Гайка', 'цинк', 'кл.пр.8', 'ГОСТ 5915-70', 'М4-М16')
TAB1SaveInExcel('Гайка', 'цинк', 'кл.пр.8', 'ГОСТ 5915-70', 'М18-М36')
TAB1SaveInExcel('Гайка', 'цинк', 'кл.пр.8', 'ГОСТ 5915-70', 'М3, М42-М72')
# Пишем данные в ТАБ2 балты
TAB2SumGroupDiamMonthForCol('М6-М16', 'Болт', 'кл.пр.5.8', 'ГОСТ 7798-70')
TAB2SumGroupDiamMonthForCol('М18-М36', 'Болт', 'кл.пр.5.8', 'ГОСТ 7798-70')
TAB2SumBolT('кл.пр.5.8', 'ГОСТ 7798-70')
TAB2SumGroupDiamMonthForCol('М6-М16', 'Болт', 'кл.пр.8.8', 'ГОСТ 7798-70')
TAB2SumGroupDiamMonthForCol('М18-М36', 'Болт', 'кл.пр.8.8', 'ГОСТ 7798-70')
TAB2SumBolT('кл.пр.8.8', 'ГОСТ 7798-70')
ItogoBolt()
# Пишем данные в ТАБ2 гайки
TAB2SumGroupDiamMonthForCol('М6-М16', 'Гайка', 'кл.пр.6', 'ГОСТ 5915-70')
TAB2SumGroupDiamMonthForCol('М18-М36', 'Гайка', 'кл.пр.6', 'ГОСТ 5915-70')
TAB2SumGroupDiamMonthForCol('М3, М42-М72', 'Гайка', 'кл.пр.6', 'ГОСТ 5915-70')
TAB2SumGaika('кл.пр.6', 'ГОСТ 5915-70')
TAB2SumGroupDiamMonthForCol('М6-М16', 'Гайка', 'кл.пр.8', 'ГОСТ 5915-70')
TAB2SumGroupDiamMonthForCol('М18-М36', 'Гайка', 'кл.пр.8', 'ГОСТ 5915-70')
TAB2SumGroupDiamMonthForCol('М3, М42-М72', 'Гайка', 'кл.пр.8', 'ГОСТ 5915-70')
TAB2SumGaika('кл.пр.8', 'ГОСТ 5915-70')
ItogoGaika()
# Мостовой Крепеж
TAB2MostKrep('Болт', 'кл.пр.10.9', 'ГОСТ Р 52644-2006', 'М16-М30')
TAB2MostKrep('Гайка', 'кл.пр.10', 'ГОСТ Р 52645-2006', 'М16-М30')

lst_name_metiz = ['Болт', 'Гайка']
lst_coating = ['черный', 'цинк']
all_month_sale = SearchLastDate() - 18

HeadWrite(ws_weight)
HeadWrite(ws_tabl1)
HeadWrite(ws_tabl2)


# Для Графиков
def SearchItogoTAB2():
    for m in ws_tabl2['1']:
        if m.value == 'Итого':
            return column_index_from_string(m.column)

dict_metiz_graph = {
    'Болты': {
        'кл.пр.5.8': {'ч': [], 'ц': [], 'ч+ц': []},
        'кл.пр.8.8': {'ч': [], 'ц': [], 'ч+ц': []},
        'кл.пр.(5.8+8.8)': []},
    'Гайки': {
        'кл.пр.6': {'ч': [], 'ц': [], 'ч+ц': []},
        'кл.пр.8': {'ч': [], 'ц': [], 'ч+ц': []},
        'кл.пр.(6+8)': []},
}


def DictForGraph(bolt_or_gaika, cl_pro4, coating, stroka):
    for e, w_m in enumerate(range(7, SearchItogoTAB2())):
        monthTAB2 = ws_tabl2.cell(row=1, column=w_m).value
        if coating !=0:
            dict_metiz_graph[bolt_or_gaika][cl_pro4][coating].append([monthTAB2])
            dict_metiz_graph[bolt_or_gaika][cl_pro4][coating][e].append(ws_tabl2.cell(row=stroka, column=w_m).value)
        else:
            dict_metiz_graph[bolt_or_gaika][cl_pro4].append([monthTAB2])
            dict_metiz_graph[bolt_or_gaika][cl_pro4][e].append(ws_tabl2.cell(row=stroka, column=w_m).value)

DictForGraph('Болты', 'кл.пр.5.8', 'ч', 8)
DictForGraph('Болты', 'кл.пр.5.8', 'ц', 9)
DictForGraph('Болты', 'кл.пр.5.8', 'ч+ц', 10)
DictForGraph('Болты', 'кл.пр.8.8', 'ч', 17)
DictForGraph('Болты', 'кл.пр.8.8', 'ц', 18)
DictForGraph('Болты', 'кл.пр.8.8', 'ч+ц', 19)
DictForGraph('Болты', 'кл.пр.(5.8+8.8)', 0, 20)
DictForGraph('Гайки', 'кл.пр.6', 'ч', 30)
DictForGraph('Гайки', 'кл.пр.6', 'ц', 31)
DictForGraph('Гайки', 'кл.пр.6', 'ч+ц', 32)
DictForGraph('Гайки', 'кл.пр.8', 'ч', 42)
DictForGraph('Гайки', 'кл.пр.8', 'ц', 43)
DictForGraph('Гайки', 'кл.пр.8', 'ч+ц', 44)
DictForGraph('Гайки', 'кл.пр.(6+8)', 0, 45)


print(dict_metiz_graph)
# print(dict_metiz_graph['Болты']['кл.пр.8.8'])
# e_cell = 1
# Заполняем таблицу данными
# for sk in tonnageData:
#     for nm in lst_name_metiz:
#         for co in lst_coating:
#             for kls_pro4 in sorted(tonnageData[sk]['кг'][nm][co]):
#                 for gst in tonnageData[sk]['кг'][nm][co][kls_pro4]:
#                     for diam in sorted(tonnageData[sk]['кг'][nm][co][kls_pro4][gst]):
#                         e_cell += 1
#                         ws_weight['A' + str(e_cell)] = sk
#                         ws_weight['B' + str(e_cell)] = nm
#                         ws_weight['C' + str(e_cell)] = co
#                         ws_weight['D' + str(e_cell)] = kls_pro4
#                         ws_weight['E' + str(e_cell)] = gst
#                         ws_weight['F' + str(e_cell)] = diam
#                         for e, m_s_s in enumerate(lst_sale):
#                             try:
#                                 ws_weight[get_column_letter(e + 7) + str(e_cell)] = \
#                                 tonnageData[sk]['кг'][nm][co][kls_pro4][gst][diam][m_s_s]['weight'] / 1000
#                             except:
#                                 ws_weight[get_column_letter(e + 7) + str(e_cell)] = 0

head = ws_sale['4']
# fill_red = PatternFill(start_color='F44336', fill_type='solid')
fill_red2 = PatternFill(start_color='E57373', fill_type='solid')
fill_min = PatternFill(start_color='3D5AFE', fill_type='solid')
fill_min2 = PatternFill(start_color='8C9EFF', fill_type='solid')
fill_except = PatternFill(start_color='9C27B0', fill_type='solid')
fill_more5 = PatternFill(start_color='FFC400', fill_type='solid')  # от 5.5
fill_3btw5 = PatternFill(start_color='E040FB', fill_type='solid')  #

ws_sale['B1'] = 'Средн<3'
ws_sale['B1'].fill = fill_min
ws_sale['C1'] = 'Средн<2'
ws_sale['C1'].fill = fill_min2
ws_sale['D1'] = '2<Ср<3'
ws_sale['D1'].fill = fill_red2
ws_sale['E1'] = '3<Ср<5.5'
ws_sale['E1'].fill = fill_3btw5
ws_sale['F1'] = 'Ср>5.5'
ws_sale['F1'].fill = fill_more5


def SearchMedium():
    for med in head:
        if med.value == 'Среднее':
            return column_index_from_string(med.column)


def SearchLastDate():
    for ld in head:
        if ld.value == 'Итого':
            return column_index_from_string(ld.column)  # Значение столбцов привел к ИНТ и минусанул 1


er = []
for rows in range(5, ws_sale.max_row):
    for col in range(18, SearchLastDate()):
        x = ws_sale.cell(row=rows, column=col)
        y = ws_sale.cell(row=rows, column=SearchMedium())
        try:
            if x.value / y.value >= 5.5:
                ja4 = ws_sale[x.coordinate]
                ja4.fill = fill_more5
            elif 3 <= x.value / y.value < 5.5:
                ja4 = ws_sale[x.coordinate]
                ja4.fill = fill_3btw5
            # elif 2 <= x.value / y.value < 3:
            #     ja4 = ws[x.coordinate]
            #     ja4.fill = fill_red
            elif x.value / y.value > 2:
                ja4 = ws_sale[x.coordinate]
                ja4.fill = fill_red2
            elif x.value / y.value < 0.33:
                ja4 = ws_sale[x.coordinate]
                ja4.fill = fill_min
            elif x.value / y.value < 0.5:
                ja4 = ws_sale[x.coordinate]
                ja4.fill = fill_min2

        except:
            if x == None:
                ja4 = [x.coordinate]
                ja4.fill = fill_except

# wb_sale.save('Color+WEIGHT_Angy.xlsx')
