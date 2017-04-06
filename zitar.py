import openpyxl
import re
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles.fonts import Font
from openpyxl.styles import PatternFill
from time import time

tic = time()
print('Загружаю Excel')
wb_zitar = openpyxl.load_workbook('Zitar.xlsx')
ws_zitar = wb_zitar.get_active_sheet()
wb_smetiz = openpyxl.load_workbook('WEIGHT_Angy.xlsx')
ws_smetiz = wb_smetiz.get_sheet_by_name('TDSheet')
toc_load_excel = time()
print('Время загрузки Excel ' + str(round((toc_load_excel - tic), 2)) + ' сек')

ws_smetiz[get_column_letter(ws_smetiz.max_column + 1) + '4'] = 'ЗИТАР'

fill_zitar_ok = PatternFill(start_color='9ACD32', fill_type='solid')
ptt_size = r'(\d+)х(\d+)'
ptt_gost = r'(\d+)-(\d+)'

for sm_rows in range(5,  ws_smetiz.max_row + 1):
    ws_nomen = ws_smetiz.cell(row=sm_rows, column=1).value
    sm_name_metiz = ws_smetiz.cell(row=sm_rows, column=14).value
    sm_coating = ws_smetiz.cell(row=sm_rows, column=13).value
    sm_cl_pro4 = ws_smetiz.cell(row=sm_rows, column=12).value
    sm_gost = ws_smetiz.cell(row=sm_rows, column=15).value
    sm_length = ws_smetiz.cell(row=sm_rows, column=11).value
    # sm_diameter = ws_smetiz.cell(row=sm_rows, column=10).value  # .replace('2M' or '3M', 'М')
    for zit_rows in range(28, ws_zitar.max_row - 16):
        zit_nomen = ws_zitar.cell(row=zit_rows, column=10).value
        zit_size = re.search(ptt_size, zit_nomen)
        zit_gost = re.search(ptt_gost, zit_nomen)
        # zit_kolvo = ws_zitar.cell(row=zit_rows, column=26).value
        zit_price = ws_zitar.cell(row=zit_rows, column=33).value
        if zit_nomen and zit_gost and zit_size:
            if 'Болт' == sm_name_metiz and 'Болт' in zit_nomen:
                if sm_coating == 'черный' and sm_cl_pro4 == 'кл.пр.5.8':
                    if zit_gost.group() in sm_gost:
                        sm_diameter = ws_smetiz.cell(row=sm_rows, column=10).value.replace('3М', '').replace('2М','').replace('М','')
                        if str(sm_length) in zit_size.groups() and sm_diameter in zit_size.groups():
                            # print(sm_diameter, sm_length,zit_size.groups(),zit_nomen, ws_nomen, sm_cl_pro4, sm_coating)
                            ws_zitar[get_column_letter(10) + str(zit_rows)].fill = fill_zitar_ok
                            # ЗАПИСАТЬ В СМЕТИЗ ПРАЙСЕ ПОСЛЕДНЮЮ КОЛОНКУ


toc_work = time()
print('Время обработки ' + str(round((toc_work - toc_load_excel), 2)) + ' сек')
print('-------------')
print('сохраняю...')
wb_zitar.save('Color-Zitar.xlsx')
wb_smetiz.save('Smetiz_Zitar.xlsx')
toc_save = time()
print('Время сохранения ' + str(round((toc_save - toc_work), 2)) + ' сек')
toc = time()
print('Полное время ' + str(round((toc - tic), 2)) + ' сек')
print('Готово, проверяй.')
